"""
Exportación de reportes a Excel y PDF.
Genera reportes comparativos entre períodos.
"""
import os
from datetime import date, datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR

VERDE = "FF92D050"
ROJO = "FFFF0000"
AMARILLO = "FFFFFF00"
AZUL_HEADER = "FF1F497D"
GRIS_CLARO = "FFD9D9D9"


def _estilo_header(ws, fila: int, col_ini: int, col_fin: int, texto: str, color=AZUL_HEADER):
    ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
    celda = ws.cell(row=fila, column=col_ini, value=texto)
    celda.font = Font(bold=True, color="FFFFFFFF", size=11)
    celda.fill = PatternFill("solid", fgColor=color)
    celda.alignment = Alignment(horizontal="center", vertical="center")


def exportar_sdi_excel(trabajadores: list[dict], patron: str = "",
                       periodo: str = "", destino: Optional[str] = None) -> str:
    """Genera reporte Excel de SDI por trabajador."""
    destino = destino or REPORTS_DIR
    wb = Workbook()
    ws = wb.active
    ws.title = "SDI Trabajadores"

    # Título
    _estilo_header(ws, 1, 1, 12, f"REPORTE SDI — {patron} — Período: {periodo}")
    ws.row_dimensions[1].height = 25

    # Subtítulo con fecha
    ws.merge_cells("A2:L2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9)

    # Encabezados
    headers = [
        "NSS", "Nombre", "Años Svc", "Días Vac", "Días Aguinaldo",
        "Prima Vac %", "SD Base", "FI", "SDI", "SDI Topado",
        "Tipo Prestaciones", "Alertas"
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF4472C4")
        c.alignment = Alignment(horizontal="center")

    # Datos
    for row_idx, t in enumerate(trabajadores, 4):
        sdi = t.get("sdi", 0)
        tope = t.get("tope_25_umas", 0)
        alerta = "SDI > 25 UMAs — topado" if sdi > tope else ""

        valores = [
            t.get("nss", ""),
            t.get("nombre", ""),
            t.get("anios_servicio", ""),
            t.get("dias_vacaciones", ""),
            t.get("dias_aguinaldo", ""),
            f"{t.get('prima_vacacional_pct', 0.25) * 100:.0f}%",
            t.get("salario_diario_base", 0),
            t.get("factor_integracion", 0),
            sdi,
            t.get("sdi_topado_imss", sdi),
            t.get("tipo_prestaciones", "ley"),
            alerta,
        ]
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            if col in (7, 8, 9, 10):
                c.number_format = '#,##0.00'
            if alerta and col == 12:
                c.fill = PatternFill("solid", fgColor=AMARILLO)
            if row_idx % 2 == 0:
                c.fill = PatternFill("solid", fgColor=GRIS_CLARO)

    # Autoajuste columnas
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].auto_size = True
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width or 0, 12
        )

    nombre = f"SDI_{patron}_{periodo}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    ruta = os.path.join(destino, nombre)
    wb.save(ruta)
    return ruta


def exportar_comparativo_excel(comparativo: dict, patron: str = "",
                               destino: Optional[str] = None) -> str:
    """Genera reporte comparativo de dos períodos."""
    destino = destino or REPORTS_DIR
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo Períodos"

    _estilo_header(ws, 1, 1, 6, f"COMPARATIVO PERÍODOS — {patron}")
    ws.row_dimensions[1].height = 25

    periodo_act = comparativo.get("periodo_actual", "")
    periodo_ant = comparativo.get("periodo_anterior", "")

    headers = ["Concepto", periodo_ant, periodo_act, "Variación $", "Variación %", "Estado"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF4472C4")
        c.alignment = Alignment(horizontal="center")

    conceptos = [
        ("Trabajadores Activos", "trabajadores_anterior", "trabajadores_actual", "var_trabajadores", None),
        ("Masa Salarial (SDI)", "masa_salarial_anterior", "masa_salarial_actual", "var_masa_salarial", "var_masa_pct"),
        ("SDI Promedio", None, "sdi_promedio_actual", None, None),
    ]

    for row_idx, (nombre, k_ant, k_act, k_var, k_pct) in enumerate(conceptos, 3):
        ws.cell(row=row_idx, column=1, value=nombre).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=comparativo.get(k_ant) if k_ant else "—")
        ws.cell(row=row_idx, column=3, value=comparativo.get(k_act))

        var = comparativo.get(k_var) if k_var else None
        pct = comparativo.get(k_pct) if k_pct else None
        ws.cell(row=row_idx, column=4, value=var)
        ws.cell(row=row_idx, column=5, value=f"{pct:.1f}%" if pct is not None else "—")

        estado_cell = ws.cell(row=row_idx, column=6)
        if var is not None:
            if var > 0:
                estado_cell.value = "▲ Aumento"
                estado_cell.fill = PatternFill("solid", fgColor=VERDE)
            elif var < 0:
                estado_cell.value = "▼ Disminución"
                estado_cell.fill = PatternFill("solid", fgColor=ROJO)
            else:
                estado_cell.value = "= Sin cambio"

    # Hoja de altas y bajas
    if "altas" in comparativo or "bajas" in comparativo:
        ws2 = wb.create_sheet("Altas y Bajas")
        _estilo_header(ws2, 1, 1, 2, "Movimientos entre períodos")
        ws2.cell(row=2, column=1, value="ALTAS (NSS)").font = Font(bold=True)
        ws2.cell(row=2, column=2, value="BAJAS (NSS)").font = Font(bold=True)
        altas = comparativo.get("altas", [])
        bajas = comparativo.get("bajas", [])
        for i, (a, b) in enumerate(zip(
            altas + [""] * max(0, len(bajas) - len(altas)),
            bajas + [""] * max(0, len(altas) - len(bajas))
        ), 3):
            ws2.cell(row=i, column=1, value=a)
            ws2.cell(row=i, column=2, value=b)

    nombre = f"Comparativo_{patron}_{periodo_act}_vs_{periodo_ant}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    ruta = os.path.join(destino, nombre)
    wb.save(ruta)
    return ruta


def exportar_incapacidades_excel(incapacidades: list, patron: str = "",
                                  destino: Optional[str] = None) -> str:
    """Genera reporte de incapacidades con subsidios."""
    destino = destino or REPORTS_DIR
    wb = Workbook()
    ws = wb.active
    ws.title = "Incapacidades"

    _estilo_header(ws, 1, 1, 10, f"REPORTE INCAPACIDADES — {patron}")
    headers = [
        "Folio", "NSS", "Nombre", "Tipo", "Fecha Inicio", "Fecha Fin",
        "Días Totales", "Días IMSS", "Subsidio Diario", "Subsidio Total"
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF4472C4")

    for row_idx, inc in enumerate(incapacidades, 3):
        es_rt = getattr(inc, "impacto_prima_riesgo", False)
        valores = [
            getattr(inc, "folio", ""),
            getattr(inc, "nss", ""),
            getattr(inc, "nombre_trabajador", ""),
            getattr(inc, "nombre_tipo", ""),
            str(getattr(inc, "fecha_inicio", "")),
            str(getattr(inc, "fecha_fin", "")),
            getattr(inc, "dias_totales", 0),
            getattr(inc, "dias_pagan_imss", 0),
            getattr(inc, "subsidio_diario_imss", 0),
            getattr(inc, "subsidio_total_imss", 0),
        ]
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            if col in (9, 10):
                c.number_format = '#,##0.00'
            if es_rt:
                c.fill = PatternFill("solid", fgColor="FFFFC7CE")  # Rosa para RT/EP

    nombre = f"Incapacidades_{patron}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    ruta = os.path.join(destino, nombre)
    wb.save(ruta)
    return ruta
